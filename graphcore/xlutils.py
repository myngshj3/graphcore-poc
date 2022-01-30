
import openpyxl as xl
from networkml.network import NetworkClass
from networkml.network import NetworkClassInstance
from networkml.network import ExtensibleWrappedAccessor


def xl_open_workbook(xlsfile: str, readonly=False):
    wb: xl.Workbook = xl.load_workbook(xlsfile)
    return wb

def xl_save_workbook(wb: xl.Workbook, filename):
    wb.save(filename)

def xl_new_workbook():
    wb = xl.Workbook()
    return wb

def xl_sheet(wb: xl.Workbook, sheet_name: str):
    ws = wb[sheet_name]
    return ws

def xl_sheet_names(wb: xl.Workbook):
    return wb.sheetnames

def xl_sheet_change_name(ws, name):
    ws.title = name

def xl_add_sheet(wb: xl.Workbook, name: str, pos=0):
    ws = wb.create_sheet(name, index=pos)
    return ws

def xl_new_sheet(wb: xl.Workbook, name: str):
    ws_new = wb.create_sheet(title=name)
    return ws_new

def xl_copy_sheet(wb: xl.Workbook, name: str):
    ws_copy = wb.copy_worksheet(wb[name])
    return ws_copy

def xl_remove_sheet(wb: xl.Workbook, name: str):
    wb.remove(name)

def xl_cell(ws, row, column):
    return ws.cell(row, column)

def xl_range(ws, range):
    return ws[range]

def xl_cell_value(cell):
    return cell.value

def xl_set_cell_value(cell, val):
    cell.value = val

def xl_handler_class(meta: NetworkClass):
    clazz: NetworkClassInstance = meta(meta, ("XlHandler", (), ()))
    m = ExtensibleWrappedAccessor(clazz, "attributes", clazz, lambda ao,c,eo,ca,ea:eo.dump_attributes())
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "load", None, lambda ao,c,eo,ca,ea:xl.load_workbook(ca[0]))
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "save", None, lambda ao,c,eo,ca,ea:ca[0].save())
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "new_workbook", None, lambda ao,c,eo,ca,ea:xl.Workbook())
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "sheet", None, lambda ao,c,eo,ca,ea:ca[0][ca[1]])
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "sheetnames", None, lambda ao,c,eo,ca,ea:ca[0].sheetnames)
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "change_sheetname", None, lambda ao,c,eo,ca,ea:xl_sheet_change_name(ca[0], ca[1]))
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "add_sheet", None, lambda ao,c,eo,ca,ea:ca[0].create_sheet(ca[1]))
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "new_sheet", None, lambda ao,c,eo,ca,ea:ca[0].create_sheet(ca[1]))
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "copy_sheet", None, lambda ao,c,eo,ca,ea:ca[0].copy_worksheet(ca[0][ca[1]]))
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "remove_sheet", None, lambda ao,c,eo,ca,ea:ca[0].remove(ca[1]))
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "cell", None, lambda ao,c,eo,ca,ea:ca[0].cell(ca[1],ca[2]))
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "range", None, lambda ao,c,eo,ca,ea:ca[0][ca[1]])
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "cell_value", None, lambda ao,c,eo,ca,ea:ca[0].value)
    clazz.register_method(clazz, m.signature, m, depth=0)
    m = ExtensibleWrappedAccessor(clazz, "set_cell_value", None, lambda ao,c,eo,ca,ea:xl_set_cell_value(ca[0],ca[1]))
    clazz.register_method(clazz, m.signature, m)
    return clazz